# %%
import openpyxl
import json
from config import config
from db_service import addrecordsdb
import string

# %%
def remove_punctuation(input_string):
    translator = str.maketrans('', '', string.punctuation)
    
    cleaned_string = input_string.translate(translator)
    
    return cleaned_string

# %%
def to_camel_case(input_string):
    words = input_string.split()

    camel_case_words = [words[0].lower()] + [word.capitalize() for word in words[1:]]

    return ''.join(camel_case_words)

# %%
def header_dict_util(sheet_obj):
    
    in_group = False
    header_row = None
    data_row = None
    file_name_lst = []
    header_dict = {}


    for row in sheet_obj.iter_rows(values_only=True):
        if any(row):  
            if not in_group:
                header_row = None
                data_row = None
                in_group = True
            if header_row is None:
                header_row = row
            elif data_row is None:
                data_row = row
                result_dict = dict(zip(header_row, data_row))
                file_name_lst.append(result_dict)
        elif in_group:
            in_group = False

    for i, result_dict in enumerate(file_name_lst, start=1):

        temp = {"title":result_dict["formTitle"], 
                "subTitle":result_dict["assetClassName"],
                "asset_class_code":result_dict["assetClassCode"],
                "asset_class_name":result_dict["assetClassName"],
                "plan_name":result_dict["planName"],
                "pm_title":result_dict["pmItem"].lower()
               }
        header_dict[f'Group_{i}'] = temp
    
    return(header_dict)
    

# %%
def body_dict_util(sheet_obj):
    
    in_group = False
    current_group = None
    file_counter = 0
    body_components=[]
    body_components_dict={}
    start_column = 1
    end_column = 8 


    for row in sheet_obj.iter_rows(values_only=True):
                
        row_data = row[start_column - 1:end_column]

        if any(row_data):

            temp = {
                "type": "container",
                "key": "",
                "label": "",
                "title": ""
                }

            if not in_group:

                in_group = True
                file_counter += 1

            if row_data[0] not in ['issueTitle','testTitle', 'tooltip', 'formTitle', 'assetClassName', 'assetClassCode', 'planName', 'pmItem']:
                

                formtitle = remove_punctuation(row_data[1])
                
                temp['key'] = to_camel_case(formtitle)
                temp['label'] = row_data[1]
                temp['title'] = row_data[0]
                
                # print(f'Group_{i} >> ', temp)
                
                body_components.append(temp)
                

        elif in_group:

            in_group = False
            body_components_dict[f'Group_{file_counter}'] = body_components
            body_components = []

    if body_components:
        body_components_dict[f'Group_{file_counter}'] = body_components

    return body_components_dict
        

# %%
def db_data(header_dict, body_components_dict, json_schema, company_id, db_table):
    
    for i in header_dict:
        
        json_schema['components'][0]['title'] = header_dict[i]['title']
        json_schema['components'][0]['subTitle'] = header_dict[i]['subTitle']
        json_schema['components'][1]['components'] = body_components_dict[i]
        
        container_show_condition = "test.value"
        for container_index in range(1, 5):
            json_schema['container'][container_index]['show'][0]['condition'] = container_show_condition
        
        new_json = json.dumps(json_schema)
        
        header_dict[i]['form_json'] = new_json
        header_dict[i]['company_id'] = company_id


        
        addrecordsdb(db_table, header_dict[i])


# %%
ref_json_file_path = '../src/ref_json.json'

# company_id='47aa8c4d-684a-4da2-9aad-cfbb851d3f6d' #cecco company (Prod)
company_id='22a8170a-f97c-432b-976a-28c4d7abf3ca' #EEE company (Prod)

db_table='PMItemMasterForms'

with open(ref_json_file_path, 'rb') as json_file:
    json_schema = json.load(json_file)

src_file_path = "../src/configs-v2-latest.xlsx"
save_dir = "../sheets/"


# %%
wb_obj = openpyxl.load_workbook(src_file_path)

for i in wb_obj.sheetnames:
        
    if "ch." in i:
        
        print('ch >>> ', i)

        sheet_obj = wb_obj[i]
        header_dict = header_dict_util(sheet_obj)
        body_components_dict = body_dict_util(sheet_obj)
        
        db_data(header_dict, body_components_dict, json_schema, company_id, db_table)


# %%


# %%


# %%



