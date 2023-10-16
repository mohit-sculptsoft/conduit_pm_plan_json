#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl
import os
import json
import psycopg2
from config import config


# In[2]:


def to_camel_case(input_string):
    words = input_string.split()

    camel_case_words = [words[0].lower()] + [word.capitalize() for word in words[1:]]

    return ''.join(camel_case_words)


# In[3]:


def header_dict_util(sheet_obj):
    
    in_group = False
    header_row = None
    data_row = None
    file_name_lst = []
    header_dict = {}


    for row in sheet_obj.iter_rows(values_only=True):
        if any(row):  
            if not in_group:
                header_row = None
                data_row = None
                in_group = True
            if header_row is None:
                header_row = row
            elif data_row is None:
                data_row = row
                result_dict = dict(zip(header_row, data_row))
                file_name_lst.append(result_dict)
        elif in_group:
            in_group = False

    for i, result_dict in enumerate(file_name_lst, start=1):

        temp = {"title":result_dict["formTitle"], 
                "subTitle":result_dict["assetClassName"],
                "asset_class_code":result_dict["assetClassCode"],
                "asset_class_name":result_dict["assetClassName"],
                "plan_name":result_dict["planName"],
                "pm_title":result_dict["pmItem"]
               }

        header_dict[f'Group_{i}'] = temp
    
    return(header_dict)
    


# In[4]:


def body_dict_util(sheet_obj):
    
    in_group = False
    current_group = None
    file_counter = 0
    body_components=[]
    body_components_dict={}



    for row in sheet_obj.iter_rows(values_only=True):

        if any(row):

            temp = {
                "type": "container",
                "key": "",
                "label": ""
                }

            if not in_group:

                in_group = True
                file_counter += 1

            if row[0] not in ['testTitle', 'tooltip', 'formTitle', 'assetClassName', 'assetClassCode', 'planName', 'pmItem']:

                temp['key'] = to_camel_case(row[0])
                temp['label'] = row[0]

                body_components.append(temp)


        elif in_group:

            in_group = False
            body_components_dict[f'Group_{file_counter}'] = body_components
            body_components = []

    return body_components_dict
        


# In[5]:


ref_json_file_path = '../src/ref_json.json'

with open(ref_json_file_path, 'rb') as json_file:
    json_schema = json.load(json_file)

src_file_path = "../src/configs-v2.xlsx"
save_dir = "../sheets/"


# In[6]:


wb_obj = openpyxl.load_workbook(src_file_path)


# In[7]:


sheet_obj = wb_obj['ch.11']


# In[8]:


company_id='22a8170a-f97c-432b-976a-28c4d7abf3ca'
header_dict = header_dict_util(sheet_obj)
body_components_dict = body_dict_util(sheet_obj)


# In[9]:


header_dict


# In[10]:


for i in header_dict:
    
    json_schema['components'][0]['title'] = header_dict[i]['title']
    json_schema['components'][0]['subTitle'] = header_dict[i]['subTitle']
    
    json_schema['components'][1]['components'] = body_components_dict[i]
    
    json_schema['container'][1]['show'][0]['condition'] = "test.value" 
    json_schema['container'][2]['show'][0]['condition'] = "test.value"
    json_schema['container'][3]['show'][0]['condition'] = "test.value" 
    json_schema['container'][4]['show'][0]['condition'] = "test.value"
    
    new_json = json.dumps(json_schema)
    
    header_dict[i]['form_json']=new_json
    header_dict[i]['company_id']=company_id
    
    import uuid
    pmitemmasterform_id = uuid.uuid4()
    header_dict[i]['pmitemmasterform_id'] = pmitemmasterform_id
    
    
    print(header_dict[i])
    print('')
    
    print(new_json)
    print('----------------')


# In[12]:


from db_service import addrecordsdb


# In[13]:


addrecordsdb("PMItemMasterForms", header_dict['Group_1'])


# In[ ]:




