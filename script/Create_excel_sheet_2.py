#!/usr/bin/env python
# coding: utf-8

# In[14]:


import openpyxl
import os


# In[15]:


def fetch_filename(sheet_obj):
    #create file names

    in_group = False
    header_row = None
    data_row = None
    file_name_lst = []
    file_name_dict = {}


    for row in sheet_obj.iter_rows(values_only=True):
        if any(row):  
            if not in_group:
                header_row = None
                data_row = None
                in_group = True
            if header_row is None:
                header_row = row
            elif data_row is None:
                data_row = row
                result_dict = dict(zip(header_row, data_row))
                file_name_lst.append(result_dict)
        elif in_group:
            in_group = False

    # Print the list of dictionaries
    for i, result_dict in enumerate(file_name_lst, start=1):

        filename = f"{result_dict['assetClassCode']}-{result_dict['planName']}-{result_dict['formTitle'].replace('/','-')}.xlsx"

        file_name_dict[f'Group_{i}'] = filename
    
    return file_name_dict


# In[16]:


def create_sheets(sheet_obj, file_name_dict, store_dir):

    in_group = False
    current_group = None
    file_counter = 1

    for row in sheet_obj.iter_rows(values_only=True):
        if any(row):  

            if not in_group:

                current_group = openpyxl.Workbook()
                in_group = True
                file_name = file_name_dict[f'Group_{file_counter}']
                file_counter += 1
                current_group_sheet = current_group.active

            current_group_sheet.append(row)  

        elif in_group:

            in_group = False

            current_group.save(store_dir+file_name)


# In[17]:


src_file_path = "../src/configs-v2.xlsx"
save_dir = "../sheets/"


# In[18]:


wb_obj = openpyxl.load_workbook(src_file_path)

for i in wb_obj.sheetnames:
    
    if "ch." in i:
        
        path = f'{save_dir}{i}/'
        print(path)

        if not os.path.exists(path):
            os.mkdir(path)
            print("Folder %s created!" % path)
        else:
            print("Folder %s already exists" % path)
        
        sheet_obj = wb_obj[i]
        file_name_dict = fetch_filename(sheet_obj)
        create_sheets(sheet_obj, file_name_dict, path)
            


# In[ ]:




